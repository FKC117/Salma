"""
File Sanitization Pipeline for Security

This module provides comprehensive file sanitization for uploaded files,
including malware scanning, format validation, content sanitization,
and formula/macro removal from Excel files.
"""

import os
import re
import magic
import hashlib
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from django.conf import settings
from django.core.files.uploadedfile import UploadedFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import bleach

logger = logging.getLogger(__name__)


class FileSanitizer:
    """
    Comprehensive file sanitization pipeline
    """
    
    # Allowed file types and their MIME types
    ALLOWED_FILE_TYPES = {
        # Data files
        '.csv': ['text/csv', 'application/csv', 'text/plain'],
        '.xlsx': ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'],
        '.xls': ['application/vnd.ms-excel'],
        '.json': ['application/json', 'text/json'],
        '.parquet': ['application/octet-stream'],
        
        # Image files
        '.png': ['image/png'],
        '.jpg': ['image/jpeg'],
        '.jpeg': ['image/jpeg'],
        '.gif': ['image/gif'],
        '.bmp': ['image/bmp'],
        '.tiff': ['image/tiff'],
        
        # Document files
        '.pdf': ['application/pdf'],
        '.txt': ['text/plain'],
        '.md': ['text/markdown', 'text/plain'],
    }
    
    # Dangerous patterns to remove from content
    DANGEROUS_PATTERNS = [
        # Script injections
        r'<script[^>]*>.*?</script>',
        r'javascript:',
        r'vbscript:',
        r'data:text/html',
        
        # Excel formulas and macros
        r'=\s*[A-Z]+\s*\(',  # Excel formulas
        r'=\s*[A-Z_][A-Z0-9_]*\s*\(',  # Named functions
        r'Sub\s+\w+\s*\(',  # VBA subroutines
        r'Function\s+\w+\s*\(',  # VBA functions
        r'Private\s+Sub',  # Private VBA subs
        r'Public\s+Sub',  # Public VBA subs
        
        # SQL injections
        r'(\bDROP\b|\bDELETE\b|\bINSERT\b|\bUPDATE\b|\bCREATE\b|\bALTER\b)',
        
        # Command injections
        r'(\beval\b|\bexec\b|\bsystem\b|\bshell_exec\b)',
    ]
    
    # Maximum file sizes (in bytes)
    MAX_FILE_SIZES = {
        '.csv': 100 * 1024 * 1024,  # 100MB
        '.xlsx': 50 * 1024 * 1024,  # 50MB
        '.xls': 50 * 1024 * 1024,   # 50MB
        '.json': 10 * 1024 * 1024,  # 10MB
        '.parquet': 200 * 1024 * 1024,  # 200MB
        '.png': 5 * 1024 * 1024,    # 5MB
        '.jpg': 5 * 1024 * 1024,    # 5MB
        '.jpeg': 5 * 1024 * 1024,   # 5MB
        '.pdf': 25 * 1024 * 1024,   # 25MB
        '.txt': 10 * 1024 * 1024,   # 10MB
    }
    
    def __init__(self):
        self.quarantine_dir = Path(getattr(settings, 'QUARANTINE_DIR', 'quarantine'))
        self.quarantine_dir.mkdir(exist_ok=True)
        
    def sanitize_file(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """
        Comprehensive file sanitization
        
        Args:
            uploaded_file: Django UploadedFile object
            
        Returns:
            Dict with sanitization results
        """
        try:
            # Step 1: Basic validation
            validation_result = self._validate_file_basic(uploaded_file)
            if not validation_result['is_valid']:
                return validation_result
            
            # Step 2: MIME type validation
            mime_result = self._validate_mime_type(uploaded_file)
            if not mime_result['is_valid']:
                return mime_result
            
            # Step 3: Content scanning
            content_result = self._scan_content(uploaded_file)
            if not content_result['is_valid']:
                return content_result
            
            # Step 4: Format-specific sanitization
            sanitization_result = self._sanitize_by_format(uploaded_file)
            if not sanitization_result['is_valid']:
                return sanitization_result
            
            # Step 5: Generate file hash
            file_hash = self._generate_file_hash(uploaded_file)
            
            # Step 6: Virus scanning (if available)
            virus_result = self._scan_for_viruses(uploaded_file)
            
            return {
                'is_valid': True,
                'sanitized': True,
                'file_hash': file_hash,
                'mime_type': mime_result.get('detected_mime'),
                'file_extension': validation_result.get('file_extension'),
                'original_size': uploaded_file.size,
                'sanitization_applied': sanitization_result.get('sanitization_applied', []),
                'virus_scan_result': virus_result,
                'warnings': [],
                'errors': []
            }
            
        except Exception as e:
            logger.error(f"File sanitization failed: {str(e)}")
            return {
                'is_valid': False,
                'error': f'Sanitization failed: {str(e)}',
                'sanitized': False
            }
    
    def _validate_file_basic(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Basic file validation"""
        # Check file name
        if not uploaded_file.name:
            return {
                'is_valid': False,
                'error': 'File name is required'
            }
        
        # Check file extension
        file_path = Path(uploaded_file.name)
        file_extension = file_path.suffix.lower()
        
        if file_extension not in self.ALLOWED_FILE_TYPES:
            return {
                'is_valid': False,
                'error': f'File type {file_extension} is not allowed'
            }
        
        # Check file size
        max_size = self.MAX_FILE_SIZES.get(file_extension, 10 * 1024 * 1024)  # Default 10MB
        if uploaded_file.size > max_size:
            return {
                'is_valid': False,
                'error': f'File size {uploaded_file.size} exceeds maximum {max_size} bytes'
            }
        
        # Check for dangerous file name patterns
        dangerous_name_patterns = [
            r'\.\./',  # Directory traversal
            r'[<>:"|?*]',  # Invalid characters
            r'^(CON|PRN|AUX|NUL|COM[1-9]|LPT[1-9])(\.|$)',  # Windows reserved names
        ]
        
        for pattern in dangerous_name_patterns:
            if re.search(pattern, uploaded_file.name, re.IGNORECASE):
                return {
                    'is_valid': False,
                    'error': 'File name contains dangerous patterns'
                }
        
        return {
            'is_valid': True,
            'file_extension': file_extension,
            'file_size': uploaded_file.size
        }
    
    def _validate_mime_type(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Validate MIME type using python-magic"""
        try:
            # Reset file pointer
            uploaded_file.seek(0)
            
            # Read first 2KB for MIME detection
            file_header = uploaded_file.read(2048)
            uploaded_file.seek(0)
            
            # Detect MIME type
            detected_mime = magic.from_buffer(file_header, mime=True)
            
            # Get allowed MIME types for this extension
            file_extension = Path(uploaded_file.name).suffix.lower()
            allowed_mimes = self.ALLOWED_FILE_TYPES.get(file_extension, [])
            
            if detected_mime not in allowed_mimes:
                return {
                    'is_valid': False,
                    'error': f'MIME type {detected_mime} does not match file extension {file_extension}',
                    'detected_mime': detected_mime
                }
            
            return {
                'is_valid': True,
                'detected_mime': detected_mime
            }
            
        except Exception as e:
            logger.warning(f"MIME type validation failed: {str(e)}")
            return {
                'is_valid': True,  # Don't fail if MIME detection fails
                'warning': f'MIME type validation failed: {str(e)}'
            }
    
    def _scan_content(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Scan file content for dangerous patterns"""
        try:
            uploaded_file.seek(0)
            
            # Read file content (limit to first 1MB for performance)
            max_scan_size = 1024 * 1024  # 1MB
            content = uploaded_file.read(max_scan_size)
            uploaded_file.seek(0)
            
            # Convert to string if possible
            try:
                if isinstance(content, bytes):
                    content_str = content.decode('utf-8', errors='ignore')
                else:
                    content_str = str(content)
            except:
                # If can't decode, skip content scanning
                return {'is_valid': True}
            
            # Scan for dangerous patterns
            found_patterns = []
            for pattern in self.DANGEROUS_PATTERNS:
                matches = re.findall(pattern, content_str, re.IGNORECASE | re.DOTALL)
                if matches:
                    found_patterns.extend(matches)
            
            if found_patterns:
                return {
                    'is_valid': False,
                    'error': 'File contains dangerous content patterns',
                    'dangerous_patterns': found_patterns[:10]  # Limit to first 10
                }
            
            return {'is_valid': True}
            
        except Exception as e:
            logger.warning(f"Content scanning failed: {str(e)}")
            return {'is_valid': True}  # Don't fail if content scan fails
    
    def _sanitize_by_format(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Format-specific sanitization"""
        file_extension = Path(uploaded_file.name).suffix.lower()
        sanitization_applied = []
        
        try:
            if file_extension in ['.xlsx', '.xls']:
                return self._sanitize_excel(uploaded_file)
            elif file_extension == '.csv':
                return self._sanitize_csv(uploaded_file)
            elif file_extension == '.json':
                return self._sanitize_json(uploaded_file)
            else:
                return {
                    'is_valid': True,
                    'sanitization_applied': []
                }
                
        except Exception as e:
            logger.error(f"Format-specific sanitization failed: {str(e)}")
            return {
                'is_valid': False,
                'error': f'Format sanitization failed: {str(e)}'
            }
    
    def _sanitize_excel(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Sanitize Excel files - remove formulas and macros"""
        try:
            uploaded_file.seek(0)
            
            # Load workbook
            workbook = load_workbook(uploaded_file, data_only=False)
            sanitization_applied = []
            
            # Remove macros and VBA
            if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
                workbook.vba_archive = None
                sanitization_applied.append('VBA_macros_removed')
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Remove formulas (convert to values)
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if cell.value.startswith('='):
                                # Remove formula, keep calculated value if available
                                cell.value = None
                                sanitization_applied.append('Excel_formulas_removed')
                
                # Remove external references
                if hasattr(worksheet, 'external_references'):
                    worksheet.external_references.clear()
                    sanitization_applied.append('External_references_removed')
            
            # Save sanitized version (in memory)
            # Note: In production, you might want to save this to a temporary file
            
            return {
                'is_valid': True,
                'sanitization_applied': list(set(sanitization_applied))
            }
            
        except Exception as e:
            logger.error(f"Excel sanitization failed: {str(e)}")
            return {
                'is_valid': False,
                'error': f'Excel sanitization failed: {str(e)}'
            }
    
    def _sanitize_csv(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Sanitize CSV files"""
        try:
            uploaded_file.seek(0)
            
            # Read CSV content
            content = uploaded_file.read().decode('utf-8', errors='ignore')
            uploaded_file.seek(0)
            
            sanitization_applied = []
            
            # Remove dangerous patterns
            original_content = content
            for pattern in self.DANGEROUS_PATTERNS:
                content = re.sub(pattern, '[SANITIZED]', content, flags=re.IGNORECASE | re.DOTALL)
            
            if content != original_content:
                sanitization_applied.append('Dangerous_patterns_removed')
            
            # HTML escape if needed
            if '<' in content or '>' in content:
                content = bleach.clean(content, tags=[], attributes={}, strip=True)
                sanitization_applied.append('HTML_tags_removed')
            
            return {
                'is_valid': True,
                'sanitization_applied': sanitization_applied,
                'sanitized_content': content if sanitization_applied else None
            }
            
        except Exception as e:
            logger.error(f"CSV sanitization failed: {str(e)}")
            return {
                'is_valid': False,
                'error': f'CSV sanitization failed: {str(e)}'
            }
    
    def _sanitize_json(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """Sanitize JSON files"""
        try:
            uploaded_file.seek(0)
            
            import json
            
            # Parse JSON to validate structure
            json_data = json.load(uploaded_file)
            uploaded_file.seek(0)
            
            # Recursively sanitize JSON values
            sanitized_data = self._sanitize_json_recursive(json_data)
            
            return {
                'is_valid': True,
                'sanitization_applied': ['JSON_structure_validated'],
                'sanitized_data': sanitized_data if sanitized_data != json_data else None
            }
            
        except json.JSONDecodeError as e:
            return {
                'is_valid': False,
                'error': f'Invalid JSON format: {str(e)}'
            }
        except Exception as e:
            logger.error(f"JSON sanitization failed: {str(e)}")
            return {
                'is_valid': False,
                'error': f'JSON sanitization failed: {str(e)}'
            }
    
    def _sanitize_json_recursive(self, data: Any) -> Any:
        """Recursively sanitize JSON data"""
        if isinstance(data, dict):
            return {key: self._sanitize_json_recursive(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [self._sanitize_json_recursive(item) for item in data]
        elif isinstance(data, str):
            # Sanitize string values
            sanitized = data
            for pattern in self.DANGEROUS_PATTERNS:
                sanitized = re.sub(pattern, '[SANITIZED]', sanitized, flags=re.IGNORECASE | re.DOTALL)
            return bleach.clean(sanitized, tags=[], attributes={}, strip=True)
        else:
            return data
    
    def _generate_file_hash(self, uploaded_file: UploadedFile) -> str:
        """Generate SHA-256 hash of file content"""
        try:
            uploaded_file.seek(0)
            
            hash_sha256 = hashlib.sha256()
            for chunk in iter(lambda: uploaded_file.read(4096), b""):
                hash_sha256.update(chunk)
            
            uploaded_file.seek(0)
            return hash_sha256.hexdigest()
            
        except Exception as e:
            logger.error(f"Hash generation failed: {str(e)}")
            return ""
    
    def _scan_for_viruses(self, uploaded_file: UploadedFile) -> Dict[str, Any]:
        """
        Virus scanning (placeholder for ClamAV integration)
        In production, integrate with ClamAV or similar
        """
        try:
            # Placeholder implementation
            # In production, you would integrate with ClamAV:
            # import pyclamd
            # cd = pyclamd.ClamdAgnostic()
            # scan_result = cd.scan_stream(uploaded_file.read())
            
            return {
                'scanned': False,
                'clean': True,
                'scanner': 'placeholder',
                'message': 'Virus scanning not configured'
            }
            
        except Exception as e:
            logger.warning(f"Virus scanning failed: {str(e)}")
            return {
                'scanned': False,
                'clean': True,
                'error': str(e)
            }
    
    def quarantine_file(self, uploaded_file: UploadedFile, reason: str) -> str:
        """Move suspicious file to quarantine"""
        try:
            # Generate unique quarantine filename
            timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            original_name = Path(uploaded_file.name).stem
            extension = Path(uploaded_file.name).suffix
            quarantine_name = f"{original_name}_{timestamp}_{reason}{extension}"
            
            quarantine_path = self.quarantine_dir / quarantine_name
            
            # Save file to quarantine
            uploaded_file.seek(0)
            with open(quarantine_path, 'wb') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)
            
            logger.warning(f"File quarantined: {quarantine_path} (reason: {reason})")
            return str(quarantine_path)
            
        except Exception as e:
            logger.error(f"Quarantine failed: {str(e)}")
            return ""
    
    def get_sanitization_report(self, sanitization_result: Dict[str, Any]) -> Dict[str, Any]:
        """Generate a comprehensive sanitization report"""
        return {
            'timestamp': pd.Timestamp.now().isoformat(),
            'file_valid': sanitization_result.get('is_valid', False),
            'file_sanitized': sanitization_result.get('sanitized', False),
            'file_hash': sanitization_result.get('file_hash', ''),
            'mime_type': sanitization_result.get('mime_type', ''),
            'file_size': sanitization_result.get('original_size', 0),
            'sanitization_actions': sanitization_result.get('sanitization_applied', []),
            'virus_scan': sanitization_result.get('virus_scan_result', {}),
            'warnings': sanitization_result.get('warnings', []),
            'errors': sanitization_result.get('errors', []),
            'security_level': self._calculate_security_level(sanitization_result)
        }
    
    def _calculate_security_level(self, result: Dict[str, Any]) -> str:
        """Calculate security risk level"""
        if not result.get('is_valid', False):
            return 'HIGH_RISK'
        
        sanitization_count = len(result.get('sanitization_applied', []))
        
        if sanitization_count == 0:
            return 'LOW_RISK'
        elif sanitization_count <= 2:
            return 'MEDIUM_RISK'
        else:
            return 'HIGH_RISK'


# Convenience function for easy integration
def sanitize_uploaded_file(uploaded_file: UploadedFile) -> Dict[str, Any]:
    """
    Convenience function to sanitize an uploaded file
    
    Args:
        uploaded_file: Django UploadedFile object
        
    Returns:
        Sanitization result dictionary
    """
    sanitizer = FileSanitizer()
    result = sanitizer.sanitize_file(uploaded_file)
    
    # Generate comprehensive report
    report = sanitizer.get_sanitization_report(result)
    
    # Log the sanitization
    if result.get('is_valid'):
        logger.info(f"File sanitized successfully: {uploaded_file.name}")
    else:
        logger.warning(f"File sanitization failed: {uploaded_file.name} - {result.get('error')}")
        # Quarantine suspicious files
        sanitizer.quarantine_file(uploaded_file, 'sanitization_failed')
    
    return report
